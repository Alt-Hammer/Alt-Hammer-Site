"""
convert_upgrades.py
-------------------
Ingests the tabulated Wargear Upgrades & Detachment Traits workbook into
structured `mechanics` (Effects) + `selection` blocks and MERGES them onto the
existing faction-wargear / detachment-trait JSON produced by the prose
converters (convert_faction_wargear.py / convert_detachment_traits.py).

This is the single mechanical-effects ingest for the Rules-Depth increment: it
supplies the data R1 (armour statlines), R3 (detachment selection rules) and R4
(trait effects) consume. It reads a deterministic table — no sentence parsing —
and HARD-ERRORS on anything ambiguous rather than silently guessing.

SOURCE
------
  ../../Countermarch 40,000 1st Edition - Wargear Upgrades and Detachment Traits by Faction.xlsx
  One sheet per faction. Layout (rows 1-based):
    row 1  : title
    row 2  : B2 = "Detachment Points Budget" value (faction-level)
    row 3  : GROUP headers  (Identity | Selection Rules | Applies To | Model | Weapon | Keyword)
    row 4  : COLUMN headers
    row 5+ : data (one row per effect; multi-row items share one ID)

OUTPUT (merged in place, mirrored to public/)
---------------------------------------------
  src/data/faction-wargear/{slug}.json     wargearItems[].mechanics.rows[]
  src/data/detachment-traits/{slug}.json   detachmentPoints + traits[].{selection,mechanics}
  public/data/... (mirror)

USAGE
-----
  python scripts/convert_upgrades.py
  python scripts/convert_upgrades.py --faction adeptus-astartes
"""

import sys
import os
import re
import json
import shutil
import argparse

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

import openpyxl
from openpyxl.utils import get_column_letter
from ah_converter_utils import slugify, ensure_dir

# ── Path configuration ────────────────────────────────────────────────────────

# Source document paths are centralised in source_paths.py — see that file
# to change where the documents live or what they are called.
from source_paths import UPGRADES_XLSX as WORKBOOK

UNITS_DIR          = os.path.join("src", "data", "units")
FACTION_WG_DIR     = os.path.join("src", "data", "faction-wargear")
DETACH_DIR         = os.path.join("src", "data", "detachment-traits")
CATALOG_DIR        = os.path.join("src", "data", "upgrade-catalogs")
PUBLIC_FACTION_WG  = os.path.join("public", "data", "faction-wargear")
PUBLIC_DETACH      = os.path.join("public", "data", "detachment-traits")
PUBLIC_CATALOG     = os.path.join("public", "data", "upgrade-catalogs")

# ── Statline column → engine char maps ────────────────────────────────────────
# Model and Weapon groups both contain "Attacks"/"Strength" headers, so we map
# them by their group (column range), never by bare header name.

MODEL_STAT_CHAR = {
    'Activation Points': 'AP', 'Movement': 'M', 'Weapon Skill': 'WS',
    'Ballistic Skill': 'BS', 'Initiative': 'I', 'Attacks': 'A', 'Strength': 'S',
    'Toughness': 'T', 'Wounds': 'W', 'Save': 'SV', 'Leadership': 'LD',
}
WEAPON_STAT_CHAR = {
    'Range': 'Rng', 'Attacks': 'A', 'Strength': 'S',
    'Armour Piercing': 'AP', 'Damage': 'D',
}
# Weapon-stat threshold tokens in Target Filter (e.g. "Strength<9").
THRESHOLD_STAT_CHAR = {
    'Strength': 'S', 'Attacks': 'A', 'Damage': 'D', 'Range': 'Rng',
    'Toughness': 'T', 'Armour Piercing': 'AP',
}

# Model-stat words usable as a *catalog* tier band (Target Domain 'model'). A band
# selects which tier of a catalog item (e.g. a Gift of Chaos) applies to a model,
# by that model's own statline — most commonly Toughness. Distinct from the weapon
# threshold map above: this is keyed by the model's stat, not a weapon's.
MODEL_STAT_WORD_CHAR = {
    'Activation Points': 'AP', 'Movement': 'M', 'Weapon Skill': 'WS',
    'Ballistic Skill': 'BS', 'Initiative': 'I', 'Attacks': 'A', 'Strength': 'S',
    'Toughness': 'T', 'Wounds': 'W', 'Save': 'SV', 'Leadership': 'LD',
}
# Band cell forms:  'Toughness 4:6'  (inclusive range) |  'Toughness <=3' / '>=10' / '<4' / '>9'
BAND_RANGE_RE = re.compile(r'^\s*([A-Za-z][A-Za-z ]*?)\s+(\d+)\s*:\s*(\d+)\s*$')
BAND_CMP_RE   = re.compile(r'^\s*([A-Za-z][A-Za-z ]*?)\s*(>=|<=|>|<)\s*(\d+)\s*$')


def parse_model_band(cell, sheet, row):
    """Parse a catalog tier's model-domain Target Filter into an inclusive band
    {char, min, max} (min/max None = open end). Hard-errors on an unknown stat,
    an inverted range, or an unrecognised form."""
    s = str(cell).strip()
    m = BAND_RANGE_RE.match(s)
    if m:
        stat, lo, hi = m.group(1).strip(), int(m.group(2)), int(m.group(3))
        if stat not in MODEL_STAT_WORD_CHAR:
            _die(sheet, row, f"catalog band on unknown model stat {stat!r} ({cell!r})")
        if lo > hi:
            _die(sheet, row, f"catalog band range inverted ({cell!r}): {lo} > {hi}")
        return {'char': MODEL_STAT_WORD_CHAR[stat], 'min': lo, 'max': hi}
    m = BAND_CMP_RE.match(s)
    if m:
        stat, op, n = m.group(1).strip(), m.group(2), int(m.group(3))
        if stat not in MODEL_STAT_WORD_CHAR:
            _die(sheet, row, f"catalog band on unknown model stat {stat!r} ({cell!r})")
        lo = hi = None
        if   op == '<=': hi = n
        elif op == '<':  hi = n - 1
        elif op == '>=': lo = n
        elif op == '>':  lo = n + 1
        return {'char': MODEL_STAT_WORD_CHAR[stat], 'min': lo, 'max': hi}
    _die(sheet, row, f"catalog Target Filter {cell!r} is not a band "
                     f"(expected 'Stat N:M' or 'Stat <=/>=/</> N')")

# The Selection Rules group is authored as 'Detachment Selection Rules' (current
# sheets) or 'Selection Rules' (earlier sheets) — accept both. 'Wargear Quantity
# Rules' (Unit Cap Group / Unit Cap) caps how many of a wargear group a unit may
# take (e.g. armour-mount cap 1 = one mount/armour per unit).
SELECTION_GROUP_NAMES = {'Selection Rules', 'Detachment Selection Rules'}
QUANTITY_GROUP_NAME = 'Wargear Quantity Rules'
GROUP_HEADERS = SELECTION_GROUP_NAMES | {
    QUANTITY_GROUP_NAME, 'Applies To Requirements',
    'Model Statline Effects', 'Weapon Statline Effects', 'Keyword Effects',
}
IDENTITY_HEADERS = {'ID', 'Category', 'Subcategory', 'Name', 'Points Cost'}

WEAPON_CLASS_TOKENS = {'Ranged', 'Melee'}
COMPARATORS = ('>=', '<=', '>', '<')

# A stat cell string that is a *set* (carries a unit/suffix) rather than a delta:
#   12"  (movement) | 2+ (save/leadership)
SET_VALUE_RE = re.compile(r'^\d+"$|^\d+\+$')
MULT_RE      = re.compile(r'^x(\d+(?:\.\d+)?)$', re.I)
THRESHOLD_RE = re.compile(r'^\s*([A-Za-z][A-Za-z ]*?)\s*(>=|<=|>|<)\s*(\d+)\s*$')

# ── Cross-source ID drift (workbook short-ID → prose-derived entry ID) ─────────
# The join tries workbook ID, then slug(Name); this map is the last resort for a
# residual, documented mismatch that neither resolves. Keep it tiny and visible —
# it is NOT a fuzzy fallback. Currently empty: all AA IDs resolve directly (short
# workbook IDs like "1st-comp" join via slug(Name); "adepts-of-the-codex" now
# matches after the Faction Rules Index name was aligned to the workbook).
ID_ALIASES = {}


class ConversionError(Exception):
    """Raised on any ambiguous/unknown token — stops the build deterministically."""


def _die(sheet, row, msg):
    where = f"[{sheet}]" + (f" row {row}" if row else "")
    raise ConversionError(f"{where}: {msg}")


# ── Weapon vocabulary (for the Target-Filter classifier) ──────────────────────

_PARAM_RE = re.compile(r'^(\d+"?|\d+\+|D\d+(?:\+\d+)?|X)$')


def _keyword_family(kw: str) -> str:
    """Strip a trailing parameter token: 'Melta 2' -> 'Melta', 'Rapid Fire 1' -> 'Rapid Fire'."""
    toks = kw.split()
    if len(toks) >= 2 and _PARAM_RE.match(toks[-1]):
        return ' '.join(toks[:-1])
    return kw


def load_weapon_vocab(faction_slug: str):
    """Return (names:set, keyword_vocab:set) for a faction's weapons.

    keyword_vocab holds both exact keywords ('Torrent', 'Melta 2') and their
    families ('Melta') so a bare 'Melta' filter token classifies as a keyword.
    """
    path = os.path.join(UNITS_DIR, f"{faction_slug}.json")
    if not os.path.exists(path):
        raise ConversionError(
            f"[{faction_slug}] units data not found ({path}); needed to classify "
            f"Target Filter tokens. Run the unit/weapon converters first.")
    data = json.load(open(path, encoding='utf-8'))
    names, exact = set(), set()

    def _collect(w):
        for k in (w.get('keywords') or []):
            exact.add(k.strip())

    for w in data.get('weapons', []):
        names.add(w['name'].strip())
        _collect(w)
        for p in (w.get('profiles') or []):
            _collect(p)

    vocab = exact | {_keyword_family(k) for k in exact}
    return names, vocab


def load_unit_vocab(faction_slug: str):
    """Return (unit_names, unit_keywords) — both lower-cased — for validating the
    eligibility on Choice Group rows.

    Ordinary eligibility fails permissive: a token matching nothing simply never
    applies, which is the right behaviour for an invisible mechanics row. A choice
    option is different — it is rendered as a selectable row in the List Builder, so
    a typo would offer the player a pick that silently does nothing. Those we check.
    """
    path = os.path.join(UNITS_DIR, f"{faction_slug}.json")
    if not os.path.exists(path):
        raise ConversionError(
            f"[{faction_slug}] units data not found ({path}); needed to validate "
            f"Choice Group eligibility. Run the unit converters first.")
    data = json.load(open(path, encoding='utf-8'))
    names, kws = set(), set()
    for u in data.get('units', []):
        if u.get('name'):
            names.add(str(u['name']).strip().lower())
        for k in (u.get('keywords') or []):
            kws.add(str(k).strip().lower())
    return names, kws


# ── Target Filter classification (bare tokens, deterministic) ─────────────────

def classify_filter(cell, names, vocab, sheet, row):
    """Classify a Target Filter cell into {kind, values}. Hard-errors on
    unknown / ambiguous / mixed-kind tokens."""
    raw_tokens = [t.strip() for t in str(cell).split('|')]
    raw_tokens = [t for t in raw_tokens if t]
    if not raw_tokens:
        return None

    kinds = set()
    threshold_vals, class_vals, kw_vals, name_vals = [], [], [], []

    for tok in raw_tokens:
        # 1) comparator → weapon-stat threshold
        if any(c in tok for c in COMPARATORS):
            m = THRESHOLD_RE.match(tok)
            if not m:
                _die(sheet, row, f"Target Filter threshold token not parseable: {tok!r}")
            stat_word, cmp_op, num = m.group(1).strip(), m.group(2), int(m.group(3))
            if stat_word not in THRESHOLD_STAT_CHAR:
                _die(sheet, row, f"Target Filter threshold on unknown stat {stat_word!r} ({tok!r})")
            threshold_vals.append({'char': THRESHOLD_STAT_CHAR[stat_word], 'cmp': cmp_op, 'value': num})
            kinds.add('weaponStat')
            continue
        # 2) Ranged / Melee → weapon class
        if tok in WEAPON_CLASS_TOKENS:
            class_vals.append(tok)
            kinds.add('weaponClass')
            continue
        # 3/4) keyword vocab vs weapon-name list (must be unambiguous)
        in_kw, in_name = tok in vocab, tok in names
        if in_kw and in_name:
            _die(sheet, row, f"Target Filter token {tok!r} is ambiguous (matches both a "
                             f"weapon keyword and a weapon name)")
        if in_kw:
            kw_vals.append(tok); kinds.add('weaponKeyword'); continue
        if in_name:
            name_vals.append(tok); kinds.add('weaponName'); continue
        _die(sheet, row, f"Target Filter token {tok!r} matches no category "
                         f"(not a comparator, Ranged/Melee, weapon keyword or weapon name)")

    if len(kinds) != 1:
        _die(sheet, row, f"Target Filter mixes kinds {sorted(kinds)} in one cell: {cell!r}")

    kind = next(iter(kinds))
    values = {'weaponStat': threshold_vals, 'weaponClass': class_vals,
              'weaponKeyword': kw_vals, 'weaponName': name_vals}[kind]
    return {'kind': kind, 'values': values}


# ── Cell parsers ──────────────────────────────────────────────────────────────

def parse_stat_cell(char, cell, sheet, row):
    """Return {char, op, value} for one statline cell, or None if blank.
    Numeric → inc/dec delta; suffixed string (2+, 12") → set."""
    if cell is None or str(cell).strip() == '':
        return None
    if isinstance(cell, str):
        val = cell.strip()
        if not SET_VALUE_RE.match(val):
            _die(sheet, row, f"stat {char} set-value {val!r} is not a recognised set "
                             f"token (expected e.g. '2+' or '12\"')")
        return {'char': char, 'op': 'set', 'value': val}
    # numeric delta
    num = cell
    if isinstance(num, float) and num.is_integer():
        num = int(num)
    if num == 0:
        return None
    return {'char': char, 'op': 'inc' if num > 0 else 'dec', 'value': abs(num)}


def parse_points(cell, sheet, row):
    """Return {op:'delta'|'mult', value} or None."""
    if cell is None or str(cell).strip() == '':
        return None
    if isinstance(cell, str):
        m = MULT_RE.match(cell.strip())
        if not m:
            _die(sheet, row, f"Points Cost {cell!r} is neither a number nor an 'xN' multiplier")
        v = float(m.group(1))
        return {'op': 'mult', 'value': int(v) if float(v).is_integer() else v}
    num = cell
    if isinstance(num, float) and num.is_integer():
        num = int(num)
    return {'op': 'delta', 'value': num}


def parse_list(cell):
    """Comma-separated keyword cell → list (blank → [])."""
    if cell is None or str(cell).strip() == '':
        return []
    return [t.strip() for t in str(cell).split(',') if t.strip()]


def parse_eligibility(cell, sheet, row):
    """Parse eligibility as OR-of-AND groups. '|' separates OR alternatives; '+'
    separates AND terms within an alternative. Returns a list of term-lists:
      'Tacticus | Gravis'    -> [['Tacticus'], ['Gravis']]   (match either)
      'Infantry + Regiment'  -> [['Infantry', 'Regiment']]   (match both)
    Blank -> [] (match any). The engine matches if ANY group's terms are all present."""
    if cell is None or str(cell).strip() == '':
        return []
    groups = []
    for alt in str(cell).split('|'):
        terms = [t.strip() for t in alt.split('+') if t.strip()]
        if terms:
            groups.append(terms)
    return groups


# ── Header layout ─────────────────────────────────────────────────────────────

def read_layout(ws, sheet):
    """Map each data column to a semantic field using the row-3 group headers
    and row-4 column headers. Returns a dict describing where each field lives."""
    max_col = ws.max_column
    row3 = {c: (str(ws.cell(3, c).value).strip() if ws.cell(3, c).value else '') for c in range(1, max_col + 1)}
    row4 = {c: (str(ws.cell(4, c).value).strip() if ws.cell(4, c).value else '') for c in range(1, max_col + 1)}

    identity = {}            # field name -> col
    selection = {}           # header -> col
    quantity = {}            # 'Unit Cap Group'/'Unit Cap' -> col
    applies = {}             # header -> col
    model_cols = []          # [(col, char)]
    weapon_cols = []         # [(col, char)]
    kw_cols = {}             # 'Keywords Added'/'Removed' -> col

    # group boundaries from row 3
    group_at = {c: row3[c] for c in range(1, max_col + 1) if row3[c] in GROUP_HEADERS}
    starts = sorted(group_at)

    def group_range(start):
        idx = starts.index(start)
        end = starts[idx + 1] - 1 if idx + 1 < len(starts) else max_col
        return range(start, end + 1)

    for c in range(1, max_col + 1):
        if row3[c] in IDENTITY_HEADERS:
            identity[row3[c]] = c

    for start in starts:
        gname = group_at[start]
        for c in group_range(start):
            hdr = row4[c]
            if not hdr:
                continue
            if gname in SELECTION_GROUP_NAMES:
                selection[hdr] = c
            elif gname == QUANTITY_GROUP_NAME:
                quantity[hdr] = c
            elif gname == 'Applies To Requirements':
                applies[hdr] = c
            elif gname == 'Model Statline Effects':
                if hdr not in MODEL_STAT_CHAR:
                    _die(sheet, 4, f"unknown Model Statline header {hdr!r}")
                model_cols.append((c, MODEL_STAT_CHAR[hdr]))
            elif gname == 'Weapon Statline Effects':
                if hdr not in WEAPON_STAT_CHAR:
                    _die(sheet, 4, f"unknown Weapon Statline header {hdr!r}")
                weapon_cols.append((c, WEAPON_STAT_CHAR[hdr]))
            elif gname == 'Keyword Effects':
                kw_cols[hdr] = c

    # required fields present?
    for f in ('ID', 'Category', 'Subcategory', 'Name', 'Points Cost'):
        if f not in identity:
            _die(sheet, 3, f"missing identity column {f!r}")
    for f in ('Detachment Points Cost', 'Subcategory Cap', 'Exclusivity Group'):
        if f not in selection:
            _die(sheet, 4, f"missing Selection Rules column {f!r}")
    for f in ('Eligibility (Who)', 'Target Domain', 'Target Filter'):
        if f not in applies:
            _die(sheet, 4, f"missing Applies To column {f!r}")
    for f in ('Keywords Added', 'Keywords Removed'):
        if f not in kw_cols:
            _die(sheet, 4, f"missing Keyword Effects column {f!r}")

    return {'identity': identity, 'selection': selection, 'quantity': quantity,
            'applies': applies, 'model_cols': model_cols, 'weapon_cols': weapon_cols,
            'kw_cols': kw_cols}


def _cell(ws, row, col):
    v = ws.cell(row, col).value
    return v if (v is not None and str(v).strip() != '') else None


# ── Row → effect entry ────────────────────────────────────────────────────────

def _row_choice(ws, r, layout, eligibility, unit_vocab, sheet):
    """Parse a row's Choice Group / Choice Pick cells.

    A row tagged with a Choice Group is one OPTION of a player decision rather than
    an effect that always applies (Aspect Host: eight Aspect Warriors units, one of
    which gains Battleline). The option's label is the Eligibility (Who) cell — the
    thing being chosen — so no extra authoring is needed to name it.

    Returns {group, optionId, label, pick} or None. `label`/`pick` are group-level and
    get lifted onto the trait by _trait_choices, which then slims the row down.
    """
    sel = layout['selection']
    if 'Choice Group' not in sel:
        return None                                  # sheet predates the columns
    grp = _cell(ws, r, sel['Choice Group'])
    if grp is None:
        return None                                  # an ordinary, always-on row
    group = str(grp).strip()

    raw = _cell(ws, r, layout['applies']['Eligibility (Who)'])
    if raw is None:
        _die(sheet, r, f"Choice Group {group!r} row has a blank Eligibility (Who) — the "
                       f"eligibility names the option the player picks, so it is required")
    label = str(raw).strip()

    unit_names, unit_kws = unit_vocab
    for term in {t for g in eligibility for t in g}:
        if term.lower() not in unit_names and term.lower() not in unit_kws:
            _die(sheet, r, f"Choice Group {group!r} option {label!r} references {term!r}, "
                           f"which is neither a unit name nor a unit keyword in this faction "
                           f"— a choice option is shown to the player, so it must resolve")

    pick = _cell(ws, r, sel['Choice Pick']) if 'Choice Pick' in sel else None
    if pick is not None:
        if isinstance(pick, float) and pick.is_integer():
            pick = int(pick)
        if not isinstance(pick, int) or isinstance(pick, bool) or pick < 1:
            _die(sheet, r, f"Choice Pick must be a positive whole number, got {pick!r}")

    return {'group': group, 'optionId': slugify(label), 'label': label, 'pick': pick}


def build_row_entry(ws, r, layout, names, vocab, unit_vocab, sheet):
    """Parse one worksheet row into the uniform effect-row entry, validating the
    model/weapon/filter <-> domain relationship.

    A Target Filter on the *model* domain is always a statline BAND (e.g.
    'Toughness 4:6'), never a weapon filter: it selects which of an item's rows
    applies to a model — and with it that row's points tier. This is how a catalog
    tier (a Gift of Chaos) and a banded wargear upgrade (Adrenal Glands: 5 pts at
    T<=3, 15 at T>=4) are both authored."""
    ap = layout['applies']
    domain = _cell(ws, r, ap['Target Domain'])
    if domain is not None:
        domain = str(domain).strip()
        if domain not in ('model', 'weapon', 'unit'):
            _die(sheet, r, f"invalid Target Domain {domain!r} (expected model|weapon|unit)")

    filt_cell = _cell(ws, r, ap['Target Filter'])
    if domain == 'model' and filt_cell:
        filt = {'kind': 'modelBand', 'band': parse_model_band(filt_cell, sheet, r)}
    else:
        filt = classify_filter(filt_cell, names, vocab, sheet, r) if filt_cell else None
        if filt is not None and domain != 'weapon':
            _die(sheet, r, f"Target Filter present on domain {domain!r} — a filter is a "
                           f"weapon filter (domain 'weapon') or a model statline band "
                           f"(domain 'model'); a unit-domain row cannot carry one")

    model_stats = [s for (c, ch) in layout['model_cols']
                   if (s := parse_stat_cell(ch, ws.cell(r, c).value, sheet, r))]
    weapon_stats = [s for (c, ch) in layout['weapon_cols']
                    if (s := parse_stat_cell(ch, ws.cell(r, c).value, sheet, r))]
    keywords = {
        'add': parse_list(ws.cell(r, layout['kw_cols']['Keywords Added']).value),
        'remove': parse_list(ws.cell(r, layout['kw_cols']['Keywords Removed']).value),
    }

    if domain == 'weapon' and model_stats:
        _die(sheet, r, "weapon-domain row populates Model Statline columns")
    if domain in ('model', 'unit') and weapon_stats:
        _die(sheet, r, f"{domain}-domain row populates Weapon Statline columns")
    # A row may legitimately omit Target Domain only if it carries no statline/
    # keyword/filter effect (a points-only entry whose real rules stay prose,
    # e.g. Dozer Blade). Any captured effect without a domain is an authoring error.
    if domain is None and (model_stats or weapon_stats or keywords['add']
                           or keywords['remove'] or filt):
        _die(sheet, r, "row carries statline/keyword/filter effects but has no Target Domain")

    eligibility = parse_eligibility(_cell(ws, r, ap['Eligibility (Who)']), sheet, r)
    entry = {
        'target': {
            'eligibility': eligibility,
            'domain': domain,
            'filter': filt,
        },
        'modelStats': model_stats,
        'weaponStats': weapon_stats,
        'keywords': keywords,
        'points': parse_points(ws.cell(r, layout['identity']['Points Cost']).value, sheet, r),
    }
    choice = _row_choice(ws, r, layout, eligibility, unit_vocab, sheet)
    if choice:
        entry['choice'] = choice
    return entry


# ── Item / trait assembly (group rows by ID) ──────────────────────────────────

def _trait_choices(item, sheet):
    """Lift a trait's Choice Group rows into player-facing choice groups.

    Rows sharing a Choice Group are mutually-exclusive options for one decision; only
    the rows whose option the player picked apply. Returns the choices list (empty when
    the trait has none) and normalises every tagged row's `choice` down to
    {group, option} — the label and pick are group-level and live on the returned
    structure, not repeated on each row.
    """
    choices, by_group = [], {}

    for row in item['rows']:
        ch = row.get('choice')
        if not ch:
            continue
        grp = ch['group']
        if grp not in by_group:
            by_group[grp] = {'choiceId': grp, 'pick': None, 'options': []}
            choices.append(by_group[grp])
        g = by_group[grp]
        if any(o['optionId'] == ch['optionId'] for o in g['options']):
            _die(sheet, item['firstRow'],
                 f"trait {item['id']!r} choice {grp!r} lists {ch['label']!r} twice")
        g['options'].append({'optionId': ch['optionId'], 'label': ch['label']})
        if ch['pick'] is not None:
            if g['pick'] is not None and g['pick'] != ch['pick']:
                _die(sheet, item['firstRow'],
                     f"trait {item['id']!r} choice {grp!r} rows disagree on Choice Pick "
                     f"({g['pick']} vs {ch['pick']}) — it is a property of the group, so "
                     f"every row of the group must agree (or leave it blank but one)")
            g['pick'] = ch['pick']
        row['choice'] = {'group': grp, 'option': ch['optionId']}

    for g in choices:
        if g['pick'] is None:
            g['pick'] = 1                    # unauthored: "select 1 of the following"
        if len(g['options']) < 2:
            _die(sheet, item['firstRow'],
                 f"trait {item['id']!r} choice {g['choiceId']!r} has only "
                 f"{len(g['options'])} option — a choice needs at least 2")
        if g['pick'] > len(g['options']):
            _die(sheet, item['firstRow'],
                 f"trait {item['id']!r} choice {g['choiceId']!r} asks the player to pick "
                 f"{g['pick']} of only {len(g['options'])} option(s)")
    return choices


def collect_items(ws, layout, names, vocab, unit_vocab, sheet):
    """Return (wargear, traits, catalog) — each a dict[id] grouped by ID, preserving
    row order (top-most first for multi-row precedence). A Category that is neither
    'Wargear Upgrades' nor 'Detachment Traits' is a selectable upgrade CATALOG (e.g.
    'Gifts of Chaos'): its rows are tiers, it is never merge-joined onto prose (the
    per-unit entitlement lives on the unit's Upgrade Allowance), and it is emitted to
    its own file. The name is data — the converter treats any such category generically."""
    ident = layout['identity']
    sel = layout['selection']
    qty = layout['quantity']
    wargear, traits, catalog = {}, {}, {}

    for r in range(5, ws.max_row + 1):
        wb_id = _cell(ws, r, ident['ID'])
        if wb_id is None:
            continue  # spacer row
        wb_id = str(wb_id).strip()
        category = str(_cell(ws, r, ident['Category']) or '').strip()
        subcat   = str(_cell(ws, r, ident['Subcategory']) or '').strip()
        name     = str(_cell(ws, r, ident['Name']) or '').strip()

        if category == 'Wargear Upgrades':
            kind, bucket = 'wargear', wargear
        elif category == 'Detachment Traits':
            kind, bucket = 'trait', traits
        elif category:
            kind, bucket = 'catalog', catalog   # any other non-blank Category
        else:
            _die(sheet, r, f"row {wb_id!r} has a blank Category")

        # Ahead of parsing the row: a Choice Group here would otherwise trip the
        # option-shaped checks in _row_choice and report the wrong problem.
        cg = layout['selection'].get('Choice Group')
        if kind != 'trait' and cg and _cell(ws, r, cg) is not None:
            _die(sheet, r, f"Choice Group is only supported on Detachment Traits rows "
                           f"({wb_id!r} is a {kind} row)")

        entry = build_row_entry(ws, r, layout, names, vocab, unit_vocab, sheet)

        if wb_id not in bucket:
            item = {'id': wb_id, 'name': name, 'category': category,
                    'subcategory': subcat, 'rows': [], 'firstRow': r}
            if kind == 'trait':
                item['detachmentPointsCost'] = _cell(ws, r, sel['Detachment Points Cost'])
                item['subcategoryCap']       = _cell(ws, r, sel['Subcategory Cap'])
                item['exclusivityGroup']     = _cell(ws, r, sel['Exclusivity Group'])
            elif kind == 'wargear':
                # Wargear Quantity Rules: a per-unit cap on a named wargear group.
                item['unitCapGroup'] = _cell(ws, r, qty['Unit Cap Group']) if 'Unit Cap Group' in qty else None
                item['unitCap']      = _cell(ws, r, qty['Unit Cap']) if 'Unit Cap' in qty else None
            # catalog: no selection/quantity fields — each row is a tier (band + points).
            bucket[wb_id] = item
        else:
            # Consistency check for shared item-level identity across rows.
            item = bucket[wb_id]
            for field, val in (('name', name), ('category', category), ('subcategory', subcat)):
                if item[field] != val:
                    _die(sheet, r, f"ID {wb_id!r} row disagrees on {field}: "
                                   f"{item[field]!r} vs {val!r}")
            if kind == 'trait':
                # DP cost / cap / exclusivity are de-duplicated: must be identical per ID.
                for field, col in (('detachmentPointsCost', 'Detachment Points Cost'),
                                   ('subcategoryCap', 'Subcategory Cap'),
                                   ('exclusivityGroup', 'Exclusivity Group')):
                    if item[field] != _cell(ws, r, sel[col]):
                        _die(sheet, r, f"trait {wb_id!r} rows disagree on {field}")
            elif kind == 'wargear':
                # Unit-cap group/cap are item-level: must be identical across rows.
                for field, hdr in (('unitCapGroup', 'Unit Cap Group'), ('unitCap', 'Unit Cap')):
                    if hdr in qty and item.get(field) != _cell(ws, r, qty[hdr]):
                        _die(sheet, r, f"wargear {wb_id!r} rows disagree on {field}")
            # catalog: rows are distinct tiers — no per-row field dedup.

        bucket[wb_id]['rows'].append(entry)

    return wargear, traits, catalog


# ── Join helpers ──────────────────────────────────────────────────────────────

def resolve_join(faction_slug, wb_id, wb_name, existing_ids, kind, sheet):
    for cand in (wb_id, slugify(wb_name)):
        if cand in existing_ids:
            return cand
    alias = ID_ALIASES.get(faction_slug, {}).get(wb_id)
    if alias:
        if alias not in existing_ids:
            _die(sheet, None, f"alias for {kind} {wb_id!r} -> {alias!r} but {alias!r} not found")
        return alias
    _die(sheet, None, f"{kind} {wb_id!r} (name {wb_name!r}) matches no existing entry "
                      f"by ID or slug(name); add an alias or fix the source ID")


# ── Merge into faction files ──────────────────────────────────────────────────

def merge_wargear(faction_slug, wargear, sheet):
    path = os.path.join(FACTION_WG_DIR, f"{faction_slug}.json")
    if not os.path.exists(path):
        raise ConversionError(f"[{faction_slug}] faction-wargear file missing ({path})")
    data = json.load(open(path, encoding='utf-8'))
    items = data.get('wargearItems', [])
    by_id = {it['itemId']: it for it in items}
    existing_ids = set(by_id)

    drift = []
    matched = 0
    for wb_id, item in wargear.items():
        target_id = resolve_join(faction_slug, wb_id, item['name'], existing_ids, 'wargear item', sheet)
        dest = by_id[target_id]
        _validate_banded_wargear(faction_slug, target_id, item['rows'])
        mech = {'rows': item['rows']}
        # Per-unit wargear cap group (e.g. armour-mount cap 1 → one mount per unit).
        if item.get('unitCapGroup') or item.get('unitCap') is not None:
            cap = item.get('unitCap')
            if isinstance(cap, float) and cap.is_integer():
                cap = int(cap)
            grp = item.get('unitCapGroup')
            mech['unitCap'] = {'group': str(grp).strip() if grp else None, 'cap': cap}
        dest['mechanics'] = mech
        matched += 1

        # Report points drift vs the prose pointsCost (workbook is authoritative;
        # we surface, we don't silently overwrite).
        pts = [row['points'] for row in item['rows'] if row['points']]
        if len(item['rows']) == 1 and pts and pts[0]['op'] == 'delta':
            wb_pts = pts[0]['value']
            if dest.get('pointsCost') is not None and dest['pointsCost'] != wb_pts:
                drift.append((target_id, dest['pointsCost'], wb_pts))

    _write(path, PUBLIC_FACTION_WG, faction_slug, data)
    return matched, drift


def merge_traits(faction_slug, traits, budget, sheet):
    path = os.path.join(DETACH_DIR, f"{faction_slug}.json")
    if not os.path.exists(path):
        raise ConversionError(f"[{faction_slug}] detachment-traits file missing ({path})")
    data = json.load(open(path, encoding='utf-8'))
    data['detachmentPoints'] = budget
    entries = data.get('detachmentTraits', [])
    by_id = {t['traitId']: t for t in entries}
    existing_ids = set(by_id)

    matched, matched_ids, with_choices = 0, set(), []
    for wb_id, item in traits.items():
        target_id = resolve_join(faction_slug, wb_id, item['name'], existing_ids, 'trait', sheet)
        dest = by_id[target_id]
        cost = item['detachmentPointsCost']
        dest['detachmentPointsCost'] = int(cost) if isinstance(cost, float) and cost.is_integer() else cost
        cap = item['subcategoryCap']
        dest['selection'] = {
            'subcategoryCap': int(cap) if isinstance(cap, float) and cap.is_integer() else cap,
            'exclusivityGroup': item['exclusivityGroup'],
        }
        # Choice groups first: it normalises the rows' `choice` keys in place.
        chs = _trait_choices(item, sheet)
        if chs:
            dest['choices'] = chs
            with_choices += [(target_id, g['choiceId'], g['pick'], len(g['options'])) for g in chs]
        else:
            dest.pop('choices', None)          # authoring withdrawn since the last run
        dest['mechanics'] = {'rows': item['rows']}
        matched += 1
        matched_ids.add(target_id)

    # Traits present in the prose file but absent from the workbook get NO selection
    # rules (DP cost / cap / exclusivity). Every selectable trait needs these even
    # when it has no permanent stat effect, so surface the gap for a source fix.
    missing = [t['traitId'] for t in entries if t['traitId'] not in matched_ids]

    _write(path, PUBLIC_DETACH, faction_slug, data)
    return matched, missing, with_choices


def _write(src_path, public_dir, faction_slug, data):
    with open(src_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    ensure_dir(public_dir)
    with open(os.path.join(public_dir, f"{faction_slug}.json"), 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ── Upgrade catalog emit (Gifts of Chaos etc.) ────────────────────────────────

def _tier_points(row, sheet):
    """A catalog tier's cost is the per-model points to add the item. The Points Cost
    column parses to a delta; a catalog tier must be a plain number (not an xN mult)."""
    p = row['points']
    if p is None:
        return 0
    if p['op'] != 'delta':
        _die(sheet, None, f"catalog tier Points Cost must be a plain number, got {p}")
    return p['value']


def _validate_bands(faction_slug, item_id, bands, what):
    """Rows keyed on a model-stat band must share one stat and not overlap, so selection
    is deterministic (each model statline falls in exactly one band). Gaps are permitted
    (a model in the gap simply gets no row / can't take the item). An item may instead be
    band-less throughout (a flat, statline-agnostic item); mixing banded and band-less
    rows is an authoring error.

    `bands` is one entry per row/tier — the band dict, or None for a band-less one."""
    banded = [b for b in bands if b]
    if not banded:
        return
    if len(banded) != len(bands):
        raise ConversionError(f"[{faction_slug}] {item_id!r} mixes banded and band-less {what}s")
    chars = {b['char'] for b in banded}
    if len(chars) != 1:
        raise ConversionError(f"[{faction_slug}] {item_id!r} mixes band stats {sorted(chars)}")
    lo = lambda b: b['min'] if b['min'] is not None else float('-inf')
    hi = lambda b: b['max'] if b['max'] is not None else float('inf')
    ordered = sorted(banded, key=lo)
    for a, b in zip(ordered, ordered[1:]):
        if lo(b) <= hi(a):
            raise ConversionError(
                f"[{faction_slug}] {item_id!r} has overlapping {what} bands "
                f"[{lo(a)}..{hi(a)}] and [{lo(b)}..{hi(b)}]")


def _model_bands(rows):
    """One entry per model-domain row: its statline band, or None when band-less."""
    out = []
    for row in rows:
        tgt = row.get('target') or {}
        if tgt.get('domain') != 'model':
            continue
        filt = tgt.get('filter') or {}
        out.append(filt.get('band') if filt.get('kind') == 'modelBand' else None)
    return out


def _validate_banded_wargear(faction_slug, item_id, rows):
    """A wargear item whose model rows are banded is priced per band (its prose
    pointsCost stays 'varies'), so each banded row must carry a plain per-model number —
    an xN multiplier has no base to multiply on the model domain."""
    bands = _model_bands(rows)
    _validate_bands(faction_slug, item_id, bands, 'row')
    if not any(bands):
        return
    for row in rows:
        tgt = row.get('target') or {}
        if tgt.get('domain') != 'model':
            continue
        pts = row.get('points')
        if pts is None:
            raise ConversionError(f"[{faction_slug}] banded wargear item {item_id!r} has a "
                                  f"band row with no Points Cost — each band carries its own cost")
        if pts['op'] != 'delta':
            raise ConversionError(f"[{faction_slug}] banded wargear item {item_id!r} band row "
                                  f"Points Cost must be a plain number, got {pts}")


def _catalog_prose(faction_slug):
    """Rules text harvested from the docx catalog H3 blocks by convert_faction_wargear,
    keyed by (category, item name) — both lower-cased. Empty when that step has not run
    (the merge is additive, so a missing file only means no prose, never a failure)."""
    path = os.path.join(FACTION_WG_DIR, f"{faction_slug}.json")
    if not os.path.exists(path):
        return {}
    data = json.load(open(path, encoding='utf-8'))
    out = {}
    for it in data.get('catalogItems', []):
        key = (str(it.get('catalogCategory', '')).strip().lower(),
               str(it.get('name', '')).strip().lower())
        if it.get('effects'):
            out[key] = (it['effects'], it.get('effectsHtml'))
    return out


def write_catalog(faction_slug, faction_name, catalog_items, sheet):
    """Emit selectable upgrade catalogs (grouped by Category, e.g. 'Gifts of Chaos') to
    their own file. No join — a catalog is stand-alone; the per-unit entitlement lives
    on each unit's Upgrade Allowance. Returns (num_catalogs, num_items)."""
    prose = _catalog_prose(faction_slug)
    by_cat, order = {}, []
    for it in catalog_items.values():
        by_cat.setdefault(it['category'], [])
        if it['category'] not in order:
            order.append(it['category'])
        by_cat[it['category']].append(it)

    catalogs, total_items = [], 0
    for cat in order:
        items = []
        for it in sorted(by_cat[cat], key=lambda x: x['firstRow']):
            tiers = []
            for row in it['rows']:
                filt = row['target']['filter']
                band = filt['band'] if (filt and filt.get('kind') == 'modelBand') else None
                tiers.append({
                    'band':        band,
                    'points':      _tier_points(row, sheet),
                    'modelStats':  row['modelStats'],
                    'weaponStats': row['weaponStats'],
                    'keywords':    row['keywords'],
                })
            _validate_bands(faction_slug, it['id'], [t['band'] for t in tiers], 'tier')
            entry = {'id': it['id'], 'name': it['name'],
                     'subcategory': it['subcategory'] or None, 'tiers': tiers}
            # Rules text from the docx (convert_faction_wargear catalogItems). A catalog
            # item's whole value can be its prose — the Asuryani Paths grant an ability
            # and a Leader expansion, several with no mechanical row at all.
            fx = prose.get((cat.strip().lower(), it['name'].strip().lower()))
            if fx:
                entry['effects'] = fx[0]
                if fx[1]:
                    entry['effectsHtml'] = fx[1]
            items.append(entry)
        catalogs.append({'category': cat, 'catalogId': slugify(cat), 'items': items})
        total_items += len(items)

    data = {'faction': faction_name, 'slug': faction_slug, 'catalogs': catalogs}
    ensure_dir(CATALOG_DIR)
    _write(os.path.join(CATALOG_DIR, f"{faction_slug}.json"), PUBLIC_CATALOG, faction_slug, data)
    return len(catalogs), total_items


# ── Main ──────────────────────────────────────────────────────────────────────

def convert_upgrades(workbook_path=WORKBOOK, faction_filter=None):
    print("")
    print("=" * 60)
    print("  Countermarch -- Converting Wargear Upgrades & Detachment Traits")
    print("=" * 60)
    print(f"  Source:  {workbook_path}")
    if faction_filter:
        print(f"  Filter:  {faction_filter}")
    print("=" * 60)
    print("")

    if not os.path.exists(workbook_path):
        raise ConversionError(f"Source workbook not found:\n     {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    processed = 0
    failed = []   # (sheet, error) — a bad sheet is ISOLATED so it can't blank the
                  # merges of every faction that follows it in the workbook. Sheets
                  # are ordered, and a single mid-list `_die` used to abort the whole
                  # run, silently leaving all later factions with no mechanics.

    for sheet in wb.sheetnames:
        faction_slug = slugify(sheet)
        if faction_filter and faction_slug != faction_filter:
            continue
        ws = wb[sheet]

        # Skip an unpopulated sheet (only title/header, no data rows).
        has_data = any(_cell(ws, r, 1) for r in range(5, ws.max_row + 1))
        if not has_data:
            print(f"  Skipping (empty): {sheet}")
            continue

        print(f"  Processing: {sheet}  ({faction_slug})")

        try:
            # Detachment Points Budget from B2.
            budget = ws.cell(2, 2).value
            if isinstance(budget, float) and budget.is_integer():
                budget = int(budget)

            names, vocab = load_weapon_vocab(faction_slug)
            unit_vocab = load_unit_vocab(faction_slug)
            layout = read_layout(ws, sheet)
            wargear, traits, catalog = collect_items(ws, layout, names, vocab, unit_vocab, sheet)

            wg_matched, drift = merge_wargear(faction_slug, wargear, sheet) if wargear else (0, [])
            tr_matched, tr_missing, tr_choices = merge_traits(faction_slug, traits, budget, sheet) if traits else (0, [], [])
            cat_n, cat_items = write_catalog(faction_slug, sheet, catalog, sheet) if catalog else (0, 0)
        except Exception as e:
            # Isolate the failure: record it, keep going. Each merge_* writes its file
            # only after fully building the data, so a mid-sheet failure leaves that
            # faction's file at its Step-5 (docx-fresh) state — never partially written.
            failed.append((sheet, f"{type(e).__name__}: {e}"))
            print(f"    ✗  SKIPPED {sheet}: {type(e).__name__}: {e}")
            print(f"       (isolated — later factions still merge; fix this sheet's row)")
            print("")
            continue

        print(f"    Detachment Points budget : {budget}")
        print(f"    Wargear items merged     : {wg_matched} (from {len(wargear)} workbook IDs)")
        print(f"    Traits merged            : {tr_matched} (from {len(traits)} workbook IDs)")
        if catalog:
            print(f"    Upgrade catalogs written : {cat_n} ({cat_items} items)")
        if tr_choices:
            print(f"    Trait sub-selections     : {len(tr_choices)} choice group(s)")
            for tid, cid, pick, n in tr_choices:
                print(f"         {tid} / {cid}: pick {pick} of {n}")
        if drift:
            print(f"    ⚠  Points drift (prose pointsCost vs workbook — workbook is authoritative):")
            for iid, prose, wbv in drift:
                print(f"         {iid}: prose {prose} vs workbook {wbv}")
        if tr_missing:
            print(f"    ⚠  Trait(s) in the Faction Rules Index with NO workbook row — they get no")
            print(f"       selection rules (DP cost / cap / exclusivity). Add a workbook row per trait:")
            for tid in tr_missing:
                print(f"         {tid}")
        processed += 1
        print("")

    print("=" * 60)
    print(f"  Complete: {processed} faction sheet(s) merged")
    if failed:
        print(f"  ✗  {len(failed)} sheet(s) failed and were skipped (their factions keep")
        print(f"     no merged mechanics until the offending row is fixed):")
        for s, err in failed:
            print(f"       {s}: {err}")
    print("=" * 60)
    print("")

    # Surface the failure so the pipeline step is flagged — but only AFTER every
    # good sheet has merged, so one bad sheet no longer blanks the rest.
    if failed:
        raise ConversionError(
            f"convert_upgrades: {len(failed)} sheet(s) failed to merge "
            f"({', '.join(s for s, _ in failed)}) — see errors above")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Ingest & merge wargear/trait mechanical effects')
    parser.add_argument('--faction', '-f', default=None, help='Slug of a single faction to process')
    parser.add_argument('--workbook', default=WORKBOOK)
    args = parser.parse_args()
    convert_upgrades(args.workbook, args.faction)
