"""
convert_units.py
────────────────
Converts the Alt-Hammer Unit Data Tables Excel file (.xlsx) into JSON
files for the website's unit stat blocks and list builder.

SOURCE FILE
───────────
Alt-Hammer 40,000 1st Edition - Unit Data Tables by Faction.xlsx
  One sheet per faction, starting at row 6 (row 4 = headers, row 5 = blank).
  Column layout (0-based indices):
    1  Unit Name
    2  Model Name
    3  Activation Points
    4  Movement
    5  Weapon Skill
    6  Ballistic Skill
    7  Initiative
    8  Attacks
    9  Strength
    10 Toughness
    11 Wounds
    12 Save
    13 Leadership
    14 Base Points per Model
    15 Squad Sizes
    16 Keywords

UNIT CATEGORIES
───────────────
Category labels (Character, Battleline, Infantry, etc.) appear as rows
where Unit Name is populated but Model Name and all stat columns are empty.
These are used to set the category for subsequent unit rows.

MULTI-MODEL UNITS
─────────────────
When consecutive rows share the same Unit Name but have different Model Names,
they are grouped under one parent unit entry with a 'models' array. The first
row's stats become the parent's stats (used for squad-level display); each
row produces a model entry with its own stat line.

OUTPUT
──────
  src/data/units/adeptus-astartes.json
  src/data/units/chaos-undivided.json
  ... etc. (one file per faction sheet)

JSON structure matches what [slug].astro consumes — see adeptus-astartes.json
for the canonical reference. The 'weapons' key is written as an empty array
by this script; convert_weapons.py populates it in a second pass.
"""

import sys
import os
import re
import json

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

import openpyxl
from ah_converter_utils import slugify, ensure_dir

# ── Path configuration ────────────────────────────────────────────────────────

UNIT_DATATABLES_XLSX = r"C:\Users\alexc\OneDrive\04 Documents\Warhammer 40k\Alt-Hammer Standalone\Alt-Hammer 40,000 1st Edition - Unit Data Tables by Faction.xlsx"

OUTPUT_DIR = "src/data/units"

# ── Sheets to skip ────────────────────────────────────────────────────────────

NON_FACTION_SHEETS = {'Hit Roll Table', 'Wound Roll Table', 'Hit Roll', 'Wound Roll'}

# ── Category labels (rows where only Unit Name is populated) ──────────────────

CATEGORY_LABELS = {
    'character', 'characters',
    'battleline',
    'infantry',
    'mounted',
    'walkers', 'walker',
    'vehicles', 'vehicle',
    'dedicated transports', 'dedicated transport',
    'fortifications', 'fortification',
    'beasts', 'beast',
    'swarms', 'swarm',
    'monsters', 'monster',
    'epic heroes', 'epic hero',
    'units',
    'aircraft',
}

# Maps sheet names to explicit slugs where the tab name doesn't match the
# faction MDX slug produced by convert_factions.py
SHEET_SLUG_OVERRIDES = {
    "The T'au Empire": 'the-tau-empire',
}

# ── Column indices (0-based) ──────────────────────────────────────────────────

COL_UNIT_NAME  = 1
COL_MODEL_NAME = 2
COL_AP         = 3
COL_M          = 4
COL_WS         = 5
COL_BS         = 6
COL_I          = 7
COL_A          = 8
COL_S          = 9
COL_T          = 10
COL_W          = 11
COL_SV         = 12
COL_LD         = 13
COL_PTS        = 14
COL_SIZES      = 15
COL_KEYWORDS   = 16   # fallback only — Keywords is now resolved by HEADER (see below),
                      # so a column inserted before it (e.g. Upgrade Allowance) can't shift it


# ── Helpers ───────────────────────────────────────────────────────────────────

def cell_val(row, col_idx, default=None):
    if col_idx < len(row):
        val = row[col_idx]
        if val is not None and str(val).strip():
            return val
    return default


def clean_str(val) -> str:
    if val is None:
        return ''
    return (str(val).strip()
        .replace('\u2018', "'").replace('\u2019', "'")
        .replace('\u201c', '"').replace('\u201d', '"')
    )


def parse_keywords(val) -> list:
    if not val:
        return []
    return [k.strip() for k in str(val).split(',') if k.strip()]


def is_category_label(unit_name: str, model_name: str, row) -> bool:
    """
    A row is a category label if:
      - unit_name is a known category keyword, AND
      - model_name is empty, AND
      - the Movement column (COL_M) is empty
    This avoids misidentifying real units whose name happens to match a label.
    """
    if unit_name.lower().strip() not in CATEGORY_LABELS:
        return False
    if model_name:
        return False
    if cell_val(row, COL_M) is not None:
        return False
    return True


# ── Squad Size Constraints (ratio caps + coupled transforms) ──────────────────
# A sparse, optional column (read by HEADER, not fixed index, so it can be
# omitted from faction sheets that don't use it, and can't be silently misread
# if column layout drifts). Two cell forms, distinguished deterministically:
#   RATIO cap : "Up to N per [X] Model"  → max = floor(count(Model)/X) * N.
#   TRANSFORM : "X ThisModel = -Y OtherModel"  → forming this model consumes Y of
#               OtherModel per X (Heavy Weapons Team: "1 Heavy Weapons Team =
#               -2 Guardsman"). Total body count is conserved.

SQUAD_CONSTRAINT_HEADER = 'Squad Size Constraints'
RATIO_RE     = re.compile(r'^(?:up\s+to\s+)?(\d+)\s+per\s+(?:(\d+)\s+)?(.+?)\s*$', re.I)
TRANSFORM_RE = re.compile(r'^(\d+)\s+(.+?)\s*=\s*-\s*(\d+)\s+(.+?)\s*$', re.I)


def find_col_by_header(ws, header_name, header_row=4):
    """Return the 0-based column index whose row-4 header equals header_name, or
    None if absent. 0-based to match cell_val()'s row-tuple indexing."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip() == header_name:
            return c - 1
    return None


# ── Upgrade Allowance (per-unit selectable-upgrade catalog entitlements) ──────
# A sparse, optional, header-driven column (same pattern as Squad Size Constraints,
# so it can be inserted anywhere / omitted per sheet without shifting fixed columns).
# It grants a unit access to a selectable upgrade catalog (e.g. the Chaos "Gifts of
# Chaos"). Authored on a unit's model row; ';'-separated allowances, each
# "<count> <scope>" where scope ∈ {unit, model, champion}:
#   'N unit'      → the unit picks N catalog items applied to EVERY model (priced per model)
#   'N model'     → EACH model of this row's type independently picks N items
#   'N champion'  → the unit's single champion model picks N extra items
# Selections are distinct within an allowance. Blank → the unit gets no catalog access.
KEYWORDS_HEADER          = 'Keywords'
UPGRADE_ALLOWANCE_HEADER = 'Upgrade Allowance'
ALLOWANCE_RE = re.compile(r'^\s*(\d+)\s+(unit|model|champion)\s*$', re.I)


def parse_upgrade_allowance(raw):
    """Parse an Upgrade Allowance cell into a list of {scope, count, distinct}, or []
    if blank. An unparseable segment yields {'_error': segment} (surfaced as a warning)."""
    if raw is None or str(raw).strip() == '':
        return []
    out = []
    for part in str(raw).split(';'):
        part = part.strip()
        if not part:
            continue
        m = ALLOWANCE_RE.match(part)
        if m:
            out.append({'scope': m.group(2).lower(), 'count': int(m.group(1)), 'distinct': True})
        else:
            out.append({'_error': part})
    return out


def _collect_allowances(raw_row, unit_name, model_type):
    """Turn one raw row's parsed allowances into unit-level entries tagged with the
    model type they were authored on. Drops (and warns about) unparseable segments."""
    out = []
    for a in raw_row.get('allowance') or []:
        if a.get('_error'):
            print(f"    ⚠  {unit_name}: unparseable Upgrade Allowance {a['_error']!r} "
                  f"on model {model_type!r} — ignored")
            continue
        out.append({'scope': a['scope'], 'count': a['count'],
                    'distinct': a['distinct'], 'modelType': model_type})
    return out


# ── Force Org Constraints (army-composition rules the List Builder enforces) ──
# A sparse, optional, header-driven column (same pattern as Squad Size Constraints
# and Upgrade Allowance). It carries the machine-readable form of a unit's "Force
# Organization" H6; the prose stays authoritative for the reader, this column for the
# validator. Authored on any one of the unit's model rows.
#
#   ';' separates atoms (all must hold)   '|' ORs references inside one atom
#   '+' ANDs terms within one reference   references are case-insensitive
#
#   max N                     flat cap per army
#   max N per <P>pts          N for every P points  ("max 5 per 500pts")
#   max N to <P>pts           tier: N while battle size is at most P
#   max N above <P>pts        tier: N once battle size exceeds P
#   max N per <ref>           N for every OTHER in-army unit matching <ref>
#   requires N <ref>          army must include N OTHER units matching <ref>
#   ignores character         exempt from the default Character cap
#   warlord                   if included, one of these must be the Warlord
#   warlord unless <ref>      …unless the army also includes a unit matching <ref>
#
# A <ref> is a unit name, a keyword, or 'Epic Hero'. Both `requires` and `max … per`
# count OTHER units: the source prose reads "at least 1 other …", and a referencing
# unit routinely carries the keyword itself (a Squadron Commander is Squadron, a
# Commissar is Regiment), which would otherwise let it satisfy its own rule.
FORCE_ORG_HEADER = 'Force Org Constraints'

_FO_PTS = r'([\d,]+)\s*(?:pts|points)'
_FO_TIER_TO    = re.compile(rf'^max\s+(\d+)\s+to\s+{_FO_PTS}$', re.I)
_FO_TIER_ABOVE = re.compile(rf'^max\s+(\d+)\s+(?:above|over)\s+{_FO_PTS}$', re.I)
_FO_PER_PTS    = re.compile(rf'^max\s+(\d+)\s+per\s+{_FO_PTS}$', re.I)
_FO_PER_REF    = re.compile(r'^max\s+(\d+)\s+per\s+(.+)$', re.I)
_FO_MAX        = re.compile(r'^max\s+(\d+)$', re.I)
_FO_REQUIRES   = re.compile(r'^requires\s+(\d+)\s+(.+)$', re.I)
_FO_IGNORES    = re.compile(r'^ignores\s+character$', re.I)
_FO_WARLORD    = re.compile(r'^warlord(?:\s+unless\s+(.+))?$', re.I)
# Tolerates the longhand OR form 'requires 1 A | requires 1 B' alongside 'requires 1 A|B'.
_FO_VERB       = re.compile(r'^(?:requires|max)\s+\d+\s+(?:per\s+)?', re.I)


def _fo_int(s):
    return int(str(s).replace(',', ''))


def _fo_refs(raw):
    """'A|B+C' → [['A'], ['B','C']] — outer list ORs, inner list ANDs."""
    alts = []
    for alt in str(raw).split('|'):
        alt = _FO_VERB.sub('', alt.strip())
        terms = [t.strip() for t in alt.split('+') if t.strip()]
        if terms:
            alts.append(terms)
    return alts


def parse_force_org_spec(raw):
    """Parse a Force Org Constraints cell into a spec with unresolved references, or
    None if blank. References stay as strings here and are resolved against the
    faction's unit names and keywords in resolve_force_org(), once the whole sheet
    has been read. Unparseable atoms collect in '_errors' (surfaced as warnings)."""
    if raw is None or str(raw).strip() == '':
        return None
    spec = {'raw': str(raw).strip()}
    for part in str(raw).split(';'):
        part = ' '.join(part.split())          # collapse whitespace, incl. around | and +
        if not part:
            continue
        m = _FO_TIER_TO.match(part)
        if m:
            spec.setdefault('tiers', []).append(
                {'maxPoints': _fo_int(m.group(2)), 'count': int(m.group(1))})
            continue
        m = _FO_TIER_ABOVE.match(part)
        if m:
            spec.setdefault('tiers', []).append(
                {'minPoints': _fo_int(m.group(2)) + 1, 'count': int(m.group(1))})
            continue
        m = _FO_PER_PTS.match(part)
        if m:
            spec['perPoints'] = {'count': int(m.group(1)), 'per': _fo_int(m.group(2))}
            continue
        m = _FO_PER_REF.match(part)
        if m:
            spec['perEach'] = {'count': int(m.group(1)), 'label': m.group(2).strip(),
                               'refs': _fo_refs(m.group(2))}
            continue
        m = _FO_MAX.match(part)
        if m:
            spec['max'] = int(m.group(1))
            continue
        m = _FO_REQUIRES.match(part)
        if m:
            spec.setdefault('requires', []).append(
                {'count': int(m.group(1)), 'label': m.group(2).strip(),
                 'refs': _fo_refs(m.group(2))})
            continue
        if _FO_IGNORES.match(part):
            spec['ignoresCharacter'] = True
            continue
        m = _FO_WARLORD.match(part)
        if m:
            spec['warlord'] = ({'unlessLabel': m.group(1).strip(),
                                'unless': _fo_refs(m.group(1))} if m.group(1) else {})
            continue
        spec.setdefault('_errors', []).append(part)
    return spec


def _unit_force_org(group, unit_name):
    """Force Org is a unit-level rule, so accept it from whichever model row carries
    it; warn if two rows of the same unit disagree."""
    found = [r['force_org'] for r in group if r.get('force_org')]
    if not found:
        return None
    for extra in found[1:]:
        if extra['raw'] != found[0]['raw']:
            print(f"    ⚠  {unit_name}: conflicting Force Org Constraints across model rows "
                  f"({found[0]['raw']!r} vs {extra['raw']!r}) — using the first")
    return found[0]


def resolve_force_org(units, sheet_name):
    """Resolve every force-org reference against this faction's unit names and keyword
    vocabulary, in place. An unresolvable reference drops its atom (with a warning)
    rather than compiling to a limit of 0 — an authoring typo must never silently
    block a legal list."""
    by_name = {u['name'].lower(): u['name'] for u in units}
    vocab = set()
    for u in units:
        for k in (u.get('keywords') or []):
            vocab.add(k.lower())
        for m in (u.get('models') or []):
            for k in (m.get('keywords') or []):
                vocab.add(k.lower())

    def resolve(alts, unit_name, label):
        out = []
        for terms in alts:
            resolved = []
            for t in terms:
                tl = t.lower()
                if tl in by_name:
                    resolved.append({'kind': 'unit', 'value': by_name[tl]})
                elif tl == 'epic hero':
                    resolved.append({'kind': 'epicHero', 'value': 'Epic Hero'})
                elif tl in vocab:
                    resolved.append({'kind': 'keyword', 'value': t})
                else:
                    print(f"    ⚠  {unit_name}: Force Org reference {t!r} in {label!r} matches no "
                          f"unit name or keyword in {sheet_name} — that rule is dropped")
                    return None
            resolved.sort(key=lambda d: (d['kind'], d['value']))
            out.append(resolved)
        return out or None

    for u in units:
        fo = u.get('forceOrg')
        if not fo:
            continue
        for bad in fo.pop('_errors', []):
            print(f"    ⚠  {u['name']}: unparseable Force Org atom {bad!r} — ignored")
        if fo.get('perEach'):
            r = resolve(fo['perEach']['refs'], u['name'], fo['perEach']['label'])
            if r:
                fo['perEach']['refs'] = r
            else:
                del fo['perEach']
        if fo.get('requires'):
            kept = []
            for rq in fo['requires']:
                r = resolve(rq['refs'], u['name'], rq['label'])
                if r:
                    rq['refs'] = r
                    kept.append(rq)
            if kept:
                fo['requires'] = kept
            else:
                del fo['requires']
        if fo.get('warlord', {}).get('unless'):
            r = resolve(fo['warlord']['unless'], u['name'], fo['warlord']['unlessLabel'])
            if r:
                fo['warlord']['unless'] = r
            else:
                fo['warlord'] = {}
        # Nothing survived but the echo of the cell — drop it so the List Builder
        # falls back to the default cascade instead of seeing an empty rule.
        if set(fo) <= {'raw'}:
            del u['forceOrg']


def parse_squad_constraint(raw):
    """Parse a Squad Size Constraints cell into a tagged dict, or None if blank.
      'Up to 2 per Shas-ui'          → {'kind':'ratio','n':2,'per':1,'model':'Shas-ui'}
      '1 Heavy Weapons Team = -2 Guardsman' → {'kind':'transform','x':1,
                                       'thisModel':'Heavy Weapons Team','y':2,'other':'Guardsman'}
    Non-blank but unparseable → {'_error': raw} (surfaced as a warning). The '='
    (transform) is checked first so it can't be mistaken for a ratio."""
    if raw is None or str(raw).strip() == '':
        return None
    s = str(raw).strip()
    mt = TRANSFORM_RE.match(s)
    if mt:
        return {'kind': 'transform', 'x': int(mt.group(1)), 'thisModel': clean_str(mt.group(2)),
                'y': int(mt.group(3)), 'other': clean_str(mt.group(4))}
    mr = RATIO_RE.match(s)
    if mr:
        return {'kind': 'ratio', 'n': int(mr.group(1)), 'per': int(mr.group(2)) if mr.group(2) else 1,
                'model': clean_str(mr.group(3))}
    return {'_error': s}


def parse_stats(row) -> dict:
    return {
        'AP':         clean_str(cell_val(row, COL_AP,     '')),
        'M':          clean_str(cell_val(row, COL_M,     '')),
        'WS':         cell_val(row, COL_WS,    None),
        'BS':         cell_val(row, COL_BS,    None),
        'I':          cell_val(row, COL_I,     None),
        'A':          cell_val(row, COL_A,     None),
        'S':          cell_val(row, COL_S,     None),
        'T':          cell_val(row, COL_T,     None),
        'W':          cell_val(row, COL_W,     None),
        'SV':         clean_str(cell_val(row, COL_SV,    '')),
        'LD':         clean_str(cell_val(row, COL_LD,    '')),
        'basePoints': cell_val(row, COL_PTS,   None),
        'squadSizes': clean_str(cell_val(row, COL_SIZES, '')),
    }


# ── Main conversion ───────────────────────────────────────────────────────────

def convert_units(xlsx_path: str, output_dir: str):
    print(f"\n{'='*60}")
    print(f"  Alt-Hammer — Converting Unit Data Tables")
    print(f"{'='*60}")
    print(f"  Source:  {xlsx_path}")
    print(f"  Output:  {output_dir}")
    print(f"{'='*60}\n")

    if not os.path.exists(xlsx_path):
        print(f"  ✗  ERROR: Source file not found:")
        print(f"     {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    written = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in NON_FACTION_SHEETS:
            print(f"  —  Skipping sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"  Processing: {sheet_name}")

        col_constraints = find_col_by_header(ws, SQUAD_CONSTRAINT_HEADER)  # None if omitted
        # Keywords is resolved by HEADER (falls back to its historical fixed index only
        # if the header is missing), so inserting Upgrade Allowance before it can't shift it.
        col_keywords = find_col_by_header(ws, KEYWORDS_HEADER)
        if col_keywords is None:
            col_keywords = COL_KEYWORDS
        col_allowance = find_col_by_header(ws, UPGRADE_ALLOWANCE_HEADER)  # None if omitted
        col_force_org = find_col_by_header(ws, FORCE_ORG_HEADER)          # None if omitted
        current_category = 'Unknown'

        # ── Pass 1: collect raw unit rows ─────────────────────────────────
        # Each entry: { unit_name, model_name, category, stats, keywords }
        raw_rows = []

        for row in ws.iter_rows(min_row=6, values_only=True):
            unit_name_val = cell_val(row, COL_UNIT_NAME)
            unit_name = clean_str(unit_name_val) if unit_name_val else ''
            if not unit_name:
                continue

            model_name_val = cell_val(row, COL_MODEL_NAME)
            model_name = clean_str(model_name_val) if model_name_val else ''

            # Category header row: unit name is a label, no model, no movement
            if is_category_label(unit_name, model_name, row):
                current_category = unit_name.title()
                continue

            # Must have a Movement value to be a valid unit stat row
            if cell_val(row, COL_M) is None:
                continue

            constraint = parse_squad_constraint(cell_val(row, col_constraints)) if col_constraints is not None else None
            allowance  = parse_upgrade_allowance(cell_val(row, col_allowance)) if col_allowance is not None else []
            force_org  = parse_force_org_spec(cell_val(row, col_force_org)) if col_force_org is not None else None
            raw_rows.append({
                'unit_name':  unit_name,
                'model_name': model_name,
                'category':   current_category,
                'stats':      parse_stats(row),
                'keywords':   parse_keywords(cell_val(row, col_keywords)),
                'constraint': constraint,
                'allowance':  allowance,
                'force_org':  force_org,
            })

        # ── Pass 2: group multi-model units ───────────────────────────────
        # Consecutive rows sharing the same unit_name are model variants.
        # Each group becomes one unit entry; if there's only one row in
        # the group, it's a single-model unit (models: null).
        units = []
        i = 0
        while i < len(raw_rows):
            row = raw_rows[i]
            unit_name = row['unit_name']

            # Collect all consecutive rows with the same unit name
            group = [row]
            j = i + 1
            while j < len(raw_rows) and raw_rows[j]['unit_name'] == unit_name:
                group.append(raw_rows[j])
                j += 1

            if len(group) == 1:
                # Single-model unit
                # Use model_name as the display name if it differs from unit_name
                # (e.g. Intercessors / Intercessor), otherwise use unit_name.
                if row.get('constraint'):
                    print(f"    ⚠  {unit_name}: Squad Size Constraint on a single-model unit "
                          f"(no other model to key off) — ignored")
                # Single-model composition names its model type after the unit.
                allowances = _collect_allowances(row, unit_name, unit_name)
                unit_obj = {
                    'name':     unit_name,
                    'category': row['category'],
                    'stats':    row['stats'],
                    'keywords': row['keywords'],
                    'models':   None,
                }
                if allowances:
                    unit_obj['upgradeAllowance'] = allowances
                force_org = _unit_force_org(group, unit_name)
                if force_org:
                    unit_obj['forceOrg'] = force_org
                units.append(unit_obj)
            else:
                # Multi-model unit — each row becomes a model entry
                model_names = {clean_str(r['model_name'] or r['unit_name']) for r in group}
                models = []
                for r in group:
                    m = {
                        'modelName': r['model_name'] or r['unit_name'],
                        'stats':     r['stats'],
                        'keywords':  r['keywords'],
                    }
                    c = r.get('constraint')
                    if c:
                        if c.get('_error'):
                            print(f"    ⚠  {unit_name}: unparseable Squad Size Constraint "
                                  f"{c['_error']!r} on model {m['modelName']!r} — ignored")
                        elif c['kind'] == 'ratio':
                            if c['model'] not in model_names:
                                print(f"    ⚠  {unit_name}: ratio driver {c['model']!r} is not a model "
                                      f"in this unit {sorted(model_names)} — ignored")
                            else:
                                m['countLimit'] = {'n': c['n'], 'per': c['per'], 'model': c['model']}
                        elif c['kind'] == 'transform':
                            if c['other'] not in model_names:
                                print(f"    ⚠  {unit_name}: transform source {c['other']!r} is not a model "
                                      f"in this unit {sorted(model_names)} — ignored")
                            elif c['y'] % c['x'] != 0:
                                print(f"    ⚠  {unit_name}: transform '{c['x']} {c['thisModel']} = -{c['y']} "
                                      f"{c['other']}' — {c['y']} not divisible by {c['x']} — ignored")
                            else:
                                m['transform'] = {'model': c['other'], 'takes': c['y'] // c['x']}
                    models.append(m)
                # Gift/upgrade allowances are authored per model row; gather them into a
                # unit-level list, each tagged with the model type it was authored on
                # (so 'model'/'champion' scopes know which model they apply to).
                allowances = []
                for r in group:
                    allowances.extend(_collect_allowances(r, unit_name, r['model_name'] or unit_name))
                unit_obj = {
                    'name':     unit_name,
                    'category': group[0]['category'],
                    'stats':    None,          # no single stat line; use models[]
                    'keywords': group[0]['keywords'],
                    'models':   models,
                }
                if allowances:
                    unit_obj['upgradeAllowance'] = allowances
                force_org = _unit_force_org(group, unit_name)
                if force_org:
                    unit_obj['forceOrg'] = force_org
                units.append(unit_obj)

            i = j

        # Force-org references can name any unit or keyword on the sheet, so they can
        # only be resolved now that every unit of the faction has been read.
        resolve_force_org(units, sheet_name)

        # ── Write JSON (weapons written by convert_weapons.py) ─────────────
        faction_slug = SHEET_SLUG_OVERRIDES.get(sheet_name, slugify(sheet_name))
        output_data = {
            'faction': sheet_name,
            'slug':    faction_slug,
            'units':   units,
            'weapons': [],     # populated by convert_weapons.py in next step
        }

        output_path = os.path.join(output_dir, f"{faction_slug}.json")
        ensure_dir(output_dir)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        print(f"    ✓  {faction_slug}.json  ({len(units)} units)")
        written += 1

    wb.close()

    print(f"\n{'='*60}")
    print(f"  Complete: {written} unit JSON files written")
    print(f"  Output directory: {output_dir}")
    print(f"  Run convert_weapons.py next to populate weapon data.")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else UNIT_DATATABLES_XLSX
    out_dir   = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_DIR
    convert_units(xlsx_path, out_dir)
