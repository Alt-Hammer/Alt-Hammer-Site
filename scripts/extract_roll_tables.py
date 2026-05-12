"""
extract_roll_tables.py
──────────────────────
Reads the Hit Roll and Wound Roll tables from the Excel source file and writes
them as structured JSON files consumed by the RollTable.astro component.

WHY THIS EXISTS
───────────────
The Hit Roll and Wound Roll tables use merged cells (a rotated "Attack Strength"
/ "Attacker's Weapons / Ballistics Skill" label spanning all rows in Column A).
python-docx cannot reliably resolve merged cell boundaries, so parsing these
tables from the Core Rules Word document produces incorrect output.

The Excel source file has the same table data in a clean, unambiguous grid
structure that reads correctly every time.

OUTPUT
──────
  src/data/hit-roll-table.json
  src/data/wound-roll-table.json

Each JSON file has this shape:
  {
    "title":    "Hit Roll Table",
    "colLabel": "Target's Weapons / Ballistics Skill",
    "rowLabel": "Attacker's Weapons / Ballistics Skill",
    "cols":     [1, 2, 3, ...],       ← column header numbers
    "rows": [
      { "label": 1, "cells": ["5+", "5+", "6+", ...] },
      { "label": 2, "cells": ["4+", "4+", "5+", ...] },
      ...
    ]
  }

HOW TO RUN
──────────
From the alt-hammer-site project folder:
  python scripts/extract_roll_tables.py

Or use run_all.py to run it as part of the full pipeline.

CONFIGURATION
─────────────
Edit the paths below if your file locations change.
"""

import os
import sys
import json

try:
    import openpyxl
except ImportError:
    print("  ✗  ERROR: openpyxl is required. Install it with:")
    print("     pip install openpyxl")
    sys.exit(1)

# ── Path configuration ────────────────────────────────────────────────────────

# Absolute path to the Hit/Wound Roll Tables Excel file
ROLL_TABLES_XLSX = (
    r"C:\Users\alexc\OneDrive\04 Documents\Warhammer 40k\Alt-Hammer Standalone"
    r"\Alt-Hammer 40,000 1st Edition - Hit and Wound Roll Tables.xlsx"
)

# Output paths — relative to where you run the script from
HIT_OUTPUT_PATH   = os.path.join("src", "data", "hit-roll-table.json")
WOUND_OUTPUT_PATH = os.path.join("src", "data", "wound-roll-table.json")

# ─────────────────────────────────────────────────────────────────────────────

# Maps sheet names to their output config
SHEET_CONFIG = {
    "Hit Roll Table": {
        "title":      "Hit Roll Table",
        "colLabel":   "Target's Weapons / Ballistics Skill",
        "rowLabel":   "Attacker's Weapons / Ballistics Skill",
        "outputPath": HIT_OUTPUT_PATH,
        # Row indices (1-based) in the Excel sheet:
        "colHeaderRow": 3,   # row containing col headers: rowLabel text + col numbers
        "dataStartRow": 4,   # first row of actual data
        # Column indices (1-based):
        "rowLabelCol":  2,   # column B — contains row number labels (1, 2, 3...)
        "dataStartCol": 3,   # column C — first column of roll value data
    },
    "Wound Roll Table": {
        "title":      "Wound Roll Table",
        "colLabel":   "Target Toughness",
        "rowLabel":   "Attack Strength",
        "outputPath": WOUND_OUTPUT_PATH,
        "colHeaderRow": 3,
        "dataStartRow": 4,
        "rowLabelCol":  2,
        "dataStartCol": 3,
    },
}


def extract_sheet(ws, config: dict) -> dict:
    """
    Parse one Excel worksheet into the JSON structure expected by RollTable.astro.
    """
    col_header_row = config["colHeaderRow"]
    data_start_row = config["dataStartRow"]
    row_label_col  = config["rowLabelCol"]
    data_start_col = config["dataStartCol"]

    # Read column header numbers from the header row
    cols = []
    for col_idx in range(data_start_col, ws.max_column + 1):
        val = ws.cell(row=col_header_row, column=col_idx).value
        if val is not None:
            cols.append(int(val))

    # Read data rows
    rows = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=row_label_col).value
        if label is None:
            continue  # skip blank rows (e.g. trailing empty rows)

        cells = []
        for col_offset in range(len(cols)):
            cell_val = ws.cell(
                row=row_idx,
                column=data_start_col + col_offset
            ).value
            # Normalise: None → 'x', everything else → lowercase string
            if cell_val is None:
                cells.append("x")
            else:
                cells.append(str(cell_val).strip().lower())

        rows.append({
            "label": int(label),
            "cells": cells,
        })

    return {
        "title":    config["title"],
        "colLabel": config["colLabel"],
        "rowLabel": config["rowLabel"],
        "cols":     cols,
        "rows":     rows,
    }


def extract_roll_tables(xlsx_path: str) -> None:
    """
    Main entry point. Reads the Excel file and writes both JSON output files.
    """
    print(f"\n{'='*60}")
    print(f"  Alt-Hammer — Extracting Roll Tables")
    print(f"{'='*60}")
    print(f"  Source:  {xlsx_path}")
    print(f"{'='*60}\n")

    if not os.path.exists(xlsx_path):
        print(f"  ✗  ERROR: Source file not found:")
        print(f"     {xlsx_path}")
        print(f"\n  Please check the ROLL_TABLES_XLSX path at the top of this script.")
        raise FileNotFoundError(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ✗  ERROR: Sheet not found in workbook: '{sheet_name}'")
            print(f"     Available sheets: {wb.sheetnames}")
            raise ValueError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        data = extract_sheet(ws, config)

        out_path = config["outputPath"]
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(
            f"  ✓  {sheet_name}: "
            f"{len(data['rows'])} rows × {len(data['cols'])} cols "
            f"→ {out_path}"
        )

    print(f"\n{'='*60}\n")


def main() -> None:
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else ROLL_TABLES_XLSX
    extract_roll_tables(xlsx_path)


if __name__ == "__main__":
    main()
