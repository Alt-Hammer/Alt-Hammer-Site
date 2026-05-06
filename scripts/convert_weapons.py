"""
convert_weapons.py
──────────────────
Converts the Alt-Hammer Weapon Data Tables Excel file (.xlsx) into JSON
and merges the result into the per-faction JSON files written by convert_units.py.

SOURCE FILE
───────────
Alt-Hammer 40,000 1st Edition - Weapon Data Tables by Faction.xlsx
  One sheet per faction, starting at row 6 (row 4 = headers, row 5 = blank).
  Column layout (0-based indices):
    1  Weapon Category   (Ranged Weapons / Melee Weapons / Psychic Attacks / Grenades)
    2  Weapon Name
    3  Weapon Attack Mode  (Normal / Supercharged / Frag / Krak / etc. — blank if single-mode)
    4  Points Cost         (populated on first profile row only; None/'-' on subsequent profiles)
    5  Range
    6  Attacks
    7  Strength
    8  Armour Piercing
    9  Damage
    10 Keywords
    11 Weapon Availability for Armour and Wargear

MULTI-MODE WEAPONS
──────────────────
Consecutive rows sharing the same Weapon Name but different Attack Mode values
are grouped as profiles under one parent weapon entry. Points come from the
first profile row; subsequent profile rows have None or '-' in the points cell.

SINGLE-MODE WEAPONS
───────────────────
Rows where Attack Mode is empty produce a flat weapon entry (profiles: null).

WEAPON CATEGORY
───────────────
Category (section) is carried forward from the last populated col 1 value.
Blank category cells inherit the previous category.

OUTPUT
──────
This script reads the existing per-faction JSON files written by convert_units.py,
populates the 'weapons' array, and writes the files back. It must run AFTER
convert_units.py. If a faction JSON file does not exist yet, it is skipped
with a warning.

  src/data/units/adeptus-astartes.json   (weapons array populated)
  src/data/units/chaos-undivided.json
  ... etc.
"""

import sys
import os
import json

script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)

import openpyxl
from ah_converter_utils import slugify, ensure_dir

# ── Path configuration ────────────────────────────────────────────────────────

WEAPON_DATATABLES_XLSX = r"C:\Users\alexc\OneDrive\04 Documents\Warhammer 40k\Alt-Hammer Standalone\Alt-Hammer 40,000 1st Edition - Weapon Data Tables by Faction.xlsx"

OUTPUT_DIR = "src/data/units"

# ── Sheets to skip ────────────────────────────────────────────────────────────

NON_FACTION_SHEETS = {'Hit Roll Table', 'Wound Roll Table', 'Hit Roll', 'Wound Roll'}

# Maps sheet names to explicit slugs
SHEET_SLUG_OVERRIDES = {
    "The T'au Empire": 'the-tau-empire',
}

# ── Column indices (0-based) ──────────────────────────────────────────────────

COL_CATEGORY   = 1
COL_WPN_NAME   = 2
COL_WPN_MODE   = 3
COL_WPN_PTS    = 4
COL_WPN_RANGE  = 5
COL_WPN_A      = 6
COL_WPN_S      = 7
COL_WPN_AP     = 8
COL_WPN_D      = 9
COL_WPN_KW     = 10
COL_WPN_AVAIL  = 11


# ── Helpers ───────────────────────────────────────────────────────────────────

def cell_val(row, col_idx, default=None):
    if col_idx < len(row):
        val = row[col_idx]
        if val is not None and str(val).strip():
            return val
    return default


def clean_str(val) -> str:
    if val is None:
        return ''
    return (str(val).strip()
        .replace('\u2018', "'").replace('\u2019', "'")
        .replace('\u201c', '"').replace('\u201d', '"')
    )


def parse_keywords(val) -> list:
    if not val:
        return []
    return [k.strip() for k in str(val).split(',') if k.strip()]


def clean_points(val):
    """
    Return points as a number if possible, '-' if it's a dash, or None.
    Subsequent profile rows often have None or '-' for points.
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s or s == '-':
        return None
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return s


# ── Main conversion ───────────────────────────────────────────────────────────

def convert_weapons(xlsx_path: str, output_dir: str):
    print(f"\n{'='*60}")
    print(f"  Alt-Hammer — Converting Weapon Data Tables")
    print(f"{'='*60}")
    print(f"  Source:  {xlsx_path}")
    print(f"  Output:  {output_dir}")
    print(f"{'='*60}\n")

    if not os.path.exists(xlsx_path):
        print(f"  ✗  ERROR: Source file not found:")
        print(f"     {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    updated = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in NON_FACTION_SHEETS:
            print(f"  —  Skipping sheet: {sheet_name}")
            continue

        faction_slug = SHEET_SLUG_OVERRIDES.get(sheet_name, slugify(sheet_name))
        json_path = os.path.join(output_dir, f"{faction_slug}.json")

        if not os.path.exists(json_path):
            print(f"  ✗  No unit JSON found for {sheet_name} at {json_path}")
            print(f"     Run convert_units.py first.")
            skipped += 1
            continue

        ws = wb[sheet_name]
        print(f"  Processing: {sheet_name}")

        current_category = 'Weapons'

        # ── Pass 1: collect all raw weapon rows ───────────────────────────
        # Each entry: { name, mode, category, points, range, attacks,
        #               strength, ap, damage, keywords, availability }
        raw_rows = []

        for row in ws.iter_rows(min_row=6, values_only=True):
            # Carry forward the category from col 1 when populated
            category_val = cell_val(row, COL_CATEGORY)
            if category_val:
                current_category = clean_str(category_val)

            wpn_name_val = cell_val(row, COL_WPN_NAME)
            wpn_name = clean_str(wpn_name_val) if wpn_name_val else ''
            if not wpn_name:
                continue

            # Must have at least one of range or attacks to be a real weapon row
            wpn_range   = clean_str(cell_val(row, COL_WPN_RANGE, ''))
            wpn_attacks = clean_str(cell_val(row, COL_WPN_A,     ''))
            if not wpn_range and not wpn_attacks:
                continue

            wpn_mode = clean_str(cell_val(row, COL_WPN_MODE, ''))

            raw_rows.append({
                'name':         wpn_name,
                'mode':         wpn_mode,       # empty string = single-mode
                'category':     current_category,
                'points':       clean_points(cell_val(row, COL_WPN_PTS, None)),
                'range':        wpn_range,
                'attacks':      wpn_attacks,
                'strength':     clean_str(cell_val(row, COL_WPN_S,  '')),
                'ap':           clean_str(cell_val(row, COL_WPN_AP, '')),
                'damage':       clean_str(cell_val(row, COL_WPN_D,  '')),
                'keywords':     parse_keywords(cell_val(row, COL_WPN_KW)),
                'availability': clean_str(cell_val(row, COL_WPN_AVAIL, '')),
            })

        # ── Pass 2: group multi-mode weapons ──────────────────────────────
        # Consecutive rows with the same weapon name and a populated mode
        # field are profiles of one weapon. Rows without a mode field are
        # single-mode weapons.
        #
        # Profile grouping rules:
        # - If a weapon has ANY row with a non-empty mode, ALL rows for that
        #   weapon name are treated as profiles.
        # - Points come from the first profile row; subsequent rows may have
        #   None (inherit) or a real value (use it).

        weapons = []
        i = 0
        while i < len(raw_rows):
            row = raw_rows[i]
            name = row['name']

            # Look ahead: collect all consecutive rows with the same name
            group = [row]
            j = i + 1
            while j < len(raw_rows) and raw_rows[j]['name'] == name:
                group.append(raw_rows[j])
                j += 1

            has_modes = any(r['mode'] for r in group)

            if len(group) == 1 and not has_modes:
                # ── Single-mode, single-row weapon ────────────────────────
                weapons.append({
                    'name':         name,
                    'section':      row['category'],
                    'points':       row['points'],
                    'profiles':     None,
                    'range':        row['range'],
                    'attacks':      row['attacks'],
                    'strength':     row['strength'],
                    'ap':           row['ap'],
                    'damage':       row['damage'],
                    'keywords':     row['keywords'],
                    'availability': row['availability'],
                })
            else:
                # ── Multi-mode weapon (profiles array) ────────────────────
                # The parent's points = first profile's points.
                # The parent's keywords and availability = first profile's values.
                # Each profile carries its own points (None if not specified,
                # meaning "same as parent / no separate cost").
                profiles = []
                parent_pts = None
                for r in group:
                    pts = r['points']
                    if parent_pts is None and pts is not None:
                        parent_pts = pts
                    profiles.append({
                        'profileName':  r['mode'] or name,
                        'points':       pts,
                        'range':        r['range'],
                        'attacks':      r['attacks'],
                        'strength':     r['strength'],
                        'ap':           r['ap'],
                        'damage':       r['damage'],
                        'keywords':     r['keywords'],
                        'availability': r['availability'],
                    })

                weapons.append({
                    'name':         name,
                    'section':      group[0]['category'],
                    'points':       parent_pts,
                    'profiles':     profiles,
                    'range':        None,
                    'attacks':      None,
                    'strength':     None,
                    'ap':           None,
                    'damage':       None,
                    'keywords':     group[0]['keywords'],
                    'availability': group[0]['availability'],
                })

            i = j

        # ── Merge into existing faction JSON ──────────────────────────────
        with open(json_path, 'r', encoding='utf-8') as f:
            faction_data = json.load(f)

        faction_data['weapons'] = weapons

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(faction_data, f, indent=2, ensure_ascii=False)

        print(f"    ✓  {faction_slug}.json  ({len(weapons)} weapons)")
        updated += 1

    wb.close()

    print(f"\n{'='*60}")
    print(f"  Complete: {updated} faction JSON files updated with weapons")
    if skipped:
        print(f"  Skipped:  {skipped} (no unit JSON found — run convert_units.py first)")
    print(f"  Output directory: {output_dir}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else WEAPON_DATATABLES_XLSX
    out_dir   = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_DIR
    convert_weapons(xlsx_path, out_dir)
